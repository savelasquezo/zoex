import os, re, requests, uuid
import hashlib

from django.utils import timezone
from django.conf import settings
from asgiref.sync import sync_to_async

from rest_framework import generics
from rest_framework import status
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl import load_workbook

from .models import UserAccount, Invoice, Withdrawals
from .serializers import UserSerializer, WithdrawalSerializer
from rest_framework.permissions import IsAuthenticated, AllowAny

from apps.core.models import Core

CONFIRMO = settings.CONFIRMO_KEY
BOLD = settings.BOLD_SECRET_KEY
APILAYER = settings.APILAYER_KEY

@sync_to_async
def AsyncUSD():
    from apps.core.models import Core
    setting = Core.objects.get(default="ZoeXConfig")
    try:
        url = "https://api.apilayer.com/fixer/latest?base=USD&symbols=COP"
        headers = {
            'apikey': f'{APILAYER}'
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            latestUSD=data["rates"]["COP"]
            setting.latestUSD = latestUSD
            setting.save()

    except Exception as e:
        eDate = timezone.now().strftime("%Y-%m-%d %H:%M")
        with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
            f.write("updateUSD {} --> Error: {}\n".format(eDate, str(e)))


def makeConfirmoInvoice(amount):
    body = {
    "product": {
        "description": "ZoexFund",
        "name": "TOPUP"
    },
    "invoice": {
        "amount": amount,
        "currencyFrom": "USD"
    },
    "settlement": {
        "currency": "USD"
    },
    "notifyEmail": "noreply@zoexbet.com",
    "notifyUrl": "https://zoexbet.com/",
    "returnUrl": "https://zoexbet.com/",
    "reference": "anything"
    }

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {CONFIRMO}'

    }
    response = requests.post('https://confirmo.net/api/v3/invoices', json=body, headers=headers)
    return response


def makeBoldInvoice(amount):
    body = {
    "product": {
        "description": "ZoeXbet",
        "name": "TOPUP"
    },
    "invoice": {
        "amount": amount,
        "currencyFrom": "USD"
    },
    "settlement": {
        "currency": "USD"
    },
    "notifyEmail": "noreply@zoexbet.com",
    "notifyUrl": "https://zoexbet.com/",
    "returnUrl": "https://zoexbet.com/",
    "reference": "anything"
    }

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {CONFIRMO}'

    }
    response = requests.post('https://confirmo.net/api/v3/invoices', json=body, headers=headers)
    return response


class fetchWithdrawals(generics.ListAPIView):
    serializer_class = WithdrawalSerializer

    def get_queryset(self):
        return Withdrawals.objects.filter(account=self.request.user)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if queryset:
            serialized_data = self.serializer_class(queryset, many=True).data
            return Response(serialized_data, status=status.HTTP_200_OK)
        else:
            return Response({'detail': 'NotFound Lottery.'}, status=status.HTTP_404_NOT_FOUND)



class requestWithdraw(generics.GenericAPIView):
    def post(self, request, *args, **kwargs):

        method = str(request.data.get('withdrawMethod', ''))
        amount = int(request.data.get('withdrawAmount', 0))

        data = {'method':method,'amount':amount}
        
        account = request.user
        if amount > account.balance or amount <= 0:
            return Response({'detail': 'The requested amount is incorrect'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            obj, created = Withdrawals.objects.update_or_create(account=account,state='pending', method=method)
            obj.amount = amount if created else obj.amount + amount
            obj.method = method
            
            currentBalance = account.balance
            account.balance = currentBalance - amount
            account.save()

            newBalance = account.balance

            apiWithdraw = str(uuid.uuid4())[:8]
            obj.voucher = apiWithdraw
            obj.save()

            file = os.path.join(settings.BASE_DIR, 'logs', 'workbook', f'{account.username}.xlsx')
            try:
                date = timezone.now().strftime("%Y-%m-%d %H:%M")
                if not os.path.exists(file):
                    WB = Workbook()
                    WS = WB.active
                    WS.append(["Tipo","Fecha","Volumen","Actual","Final","Voucher","Metodo"])
                else:
                    WB = load_workbook(file)
                    WS = WB.active

                data = [1, date, amount, currentBalance, newBalance, apiWithdraw, method]
                WS.append(data)
                WB.save(file)
                
            except Exception as e:
                with open(os.path.join(settings.BASE_DIR, 'logs/workbook.log'), 'a') as f:
                    f.write("WorkbookError: {}\n".format(str(e)))

            return Response({'apiWithdraw': apiWithdraw, 'newBalance': newBalance}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'detail': 'NotFound Lottery.'}, status=status.HTTP_404_NOT_FOUND)


class refreshInvoices(generics.GenericAPIView):
    def get(self, request, *args, **kwargs):
        try:
            account= request.user
            list_invoices_bold = Invoice.objects.filter(account=account,state="pending",method="bold")
            list_invoices_crypto = Invoice.objects.filter(account=account,state="pending",method="crypto")

            if list_invoices_crypto.exists():
                headers = {'Content-Type': 'application/json',
                            'Authorization': f'Bearer {CONFIRMO}'}
                
                for obj in list_invoices_crypto:
                    response = requests.get(f'https://confirmo.net/api/v3/invoices/{obj.voucher}', headers=headers)
                    currentStatus = response.json().get('status') if response.status_code == 200 else "pending"
                    if currentStatus == "done":
                        obj.state = currentStatus
                        obj.save()

            return Response({'detail': 'Invoices Refresh!'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': 'NotFound CryptoInvoices!'}, status=status.HTTP_404_NOT_FOUND)
        


class requestInvoice(generics.GenericAPIView):
    def post(self, request, *args, **kwargs):

        method = str(request.data.get('method', ''))
        amount = int(request.data.get('amount', 0))

        data = {'method':method,'amount':amount}
    
        try:
            obj, created = Invoice.objects.update_or_create(account=request.user,state='pending', method=method, defaults=data)
            integritySignature = "N/A"

            if method == "crypto":
                response = makeConfirmoInvoice(amount)
                if response.status_code == 201:
                    apiInvoice = response.json().get('id')
                
            if method == "bold":
                setting = Core.objects.get(default="ZoeXConfig")
                AsyncUSD()
                
                boldAmmount = int(amount*setting.latestUSD)
                apiInvoice = str(uuid.uuid4())[:12]

                hash256 = "{}{}{}{}".format(apiInvoice,str(boldAmmount),"COP",BOLD)
                m = hashlib.sha256()
                m.update(hash256.encode())
                integritySignature = m.hexdigest()

            obj.voucher = apiInvoice
            obj.save()
            return Response({'apiInvoice': apiInvoice, 'integritySignature': integritySignature, 'boldAmmount': boldAmmount}, status=status.HTTP_200_OK)
        
        except Exception as e:
            date = timezone.now().strftime("%Y-%m-%d %H:%M")
            with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
                f.write("requestInvoice {} --> Error: {}\n".format(date, str(e)))
            return Response({'error': 'NotFound Invoice.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class updateAccountInfo(generics.GenericAPIView):

    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            frame = str(request.data.get('frame', ''))
            location = str(request.data.get('location', ''))
            billing = str(request.data.get('billing', ''))
            user = UserAccount.objects.get(email=self.request.user.email)

            user.frame = frame
            user.location = location
            user.billing = billing
            user.save()

            serializer = UserSerializer(user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except Exception as e:
            eDate = timezone.now().strftime("%Y-%m-%d %H:%M")
            with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
                f.write("updateAccountInfo {} --> Error: {}\n".format(eDate, str(e)))
            return Response({'error': 'NotFound Account.'}, status=status.HTTP_404_NOT_FOUND)


        