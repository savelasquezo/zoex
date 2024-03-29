import os
from django.utils import timezone
from django.conf import settings

import pandas as pd

from django.core.mail import send_mail
from django.template.loader import render_to_string

from openpyxl import Workbook
from openpyxl import load_workbook


from django.core.mail import EmailMultiAlternatives
import base64

def sendEmailTicket(Template, requestSubject, requestEmail, requestImage, requestTicket):
    try:
        image64 = base64.b64decode(requestImage)
        email = EmailMultiAlternatives(requestSubject, '', 'noreply@zoexbet.com', [requestEmail])
        email.attach('ticket.jpg', image64, 'image/jpeg')
        email_body = render_to_string(Template, {"requestTicket": requestTicket})
        email.attach_alternative(email_body, 'text/html')
        email.send()
        
    except Exception as e:
        eDate = timezone.now().strftime('%Y-%m-%d %H:%M')
        with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
            f.write("TicketEmailError {} --> Error: {}\n".format(eDate, str(e)))


def xlsxSave(account, types, amount, balance, newBalance, credits, newCredits, id, method):
    try:
        file = os.path.join(settings.BASE_DIR, 'logs', 'xlsx', f'{account}.xlsx')
        date = timezone.now().strftime("%Y-%m-%d %H:%M")
        if not os.path.exists(file):
            WB = Workbook()
            WS = WB.active
            WS.append(["Type","Date","Ammount","Balance","newBalance","Credits","newCredits","Voucher","Method"])
        else:
            WB = load_workbook(file)
            WS = WB.active

        data = [types, date, amount, balance, newBalance, credits, newCredits, id, method]
        WS.append(data)
        WB.save(file)

    except Exception as e:
        with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
            f.write("WorkbookError: {}\n".format(str(e)))


def xlsxData(account, obj):
    try:
        df = pd.read_excel(os.path.join(settings.BASE_DIR, 'logs', 'xlsx', f'{account}.xlsx'))
        data = df[df['type'].isin([obj])]
        return data

    except Exception as e:
        with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
            f.write("xlsxData: {}\n".format(str(e)))
